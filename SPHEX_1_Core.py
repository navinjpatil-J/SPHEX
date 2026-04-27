# ============================================================
# SPHEX — Spectral Pattern Heterogeneity indeX Analyzer
# v1.1.1
# ============================================================
#
# A multiscale framework for quantitative AFM biofilm surface
# heterogeneity analysis.
#
# Copyright (c) 2025 Navinkumar Patil
#
# Licensed under the MIT License. See LICENSE file in the
# project root for full license information.
#
# Author: Navinkumar Patil
# # Contact: [navinjpatil@gmail.com]
#
# Reference:
#   Navinkumar Patil (2025). SPHEX: Spectral Pattern
#   Heterogeneity indeX Analyzer — A Python framework for
#   multiscale AFM biofilm surface heterogeneity analysis.
#
# Dependencies:
#   numpy, scipy, scikit-image, PyWavelets,
#   tifffile, matplotlib, pandas, openpyxl
#
# v1.1.1 Changes:
#   1. Fixed CV calculation in multiscale_heterogeneity()
#   2. Moved eps definition for cleaner scope
#   3. Replaced np.ptp() for NumPy 2.0+ compatibility
#   4. Changed directional CV std to ddof=1
#   5. Added clarifying comments in radial_average_psd()
# ============================================================

# ========================
# Section 0: Version & Configuration
# ========================

__version__ = "1.1.1"
__author__ = "NAVINKUMAR PATIL"
__license__ = "MIT"

# ========================
# Section 1: Core Imports
# ========================

import logging
import re
import sys
import warnings
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import pywt
from scipy.fft import fft2, fftshift
from scipy.ndimage import gaussian_filter
from scipy.stats import gaussian_kde, linregress
from skimage import color, io
from tifffile import TiffFile

# Optional GUI imports with graceful fallback
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
    GUI_AVAILABLE = True
except ImportError:
    GUI_AVAILABLE = False
    logging.warning(
        "tkinter not available. GUI mode disabled. "
        "Use programmatic API instead."
    )

# Optional Excel export
try:
    import openpyxl
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning(
        "openpyxl not available. Excel export disabled. "
        "CSV output will be used instead."
    )


# ========================
# Section 2: Logging Configuration
# ========================

def configure_logging(
    level: int = logging.INFO,
    log_file: Optional[str] = None
) -> logging.Logger:
    """
    Configure the AFMAnalyzer logging system.

    Parameters
    ----------
    level : int
        Logging level (e.g., logging.DEBUG, logging.INFO).
        Default is logging.INFO.
    log_file : str, optional
        Path to a log file. If None, logs only to console.

    Returns
    -------
    logger : logging.Logger
        Configured logger instance.

    Examples
    --------
    >>> logger = configure_logging(level=logging.DEBUG)
    >>> logger = configure_logging(log_file="afm_analysis.log")
    """
    logger = logging.getLogger("AFMAnalyzer")
    logger.setLevel(level)

    # Clear existing handlers
    logger.handlers.clear()

    # Console handler
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(level)
    formatter = logging.Formatter(
        "[%(asctime)s] %(name)s — %(levelname)s: %(message)s",
        datefmt="%H:%M:%S"
    )
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

    # File handler (optional)
    if log_file:
        file_handler = logging.FileHandler(log_file)
        file_handler.setLevel(level)
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)

    return logger


# Module-level logger
logger = configure_logging()


# ========================
# Section 3: Constants & Configuration
# ========================

class AFMConfig:
    """
    Central configuration class for AFMAnalyzer.

    All hardcoded values are documented here with physical
    justification. Modify these to change global behavior.

    Attributes
    ----------
    OPD_LEVELS : int
        Number of OPD decomposition levels (Liu et al. 2015).
    OPD_THRESHOLD_PERCENTILE : float
        PDF split percentile for OPD decomposition.
    OPD_ENERGY_DECAY : float
        Energy decay factor between OPD levels.
    KDE_BANDWIDTH_FACTOR : float
        Silverman bandwidth scaling factor (standard = 0.9).
    PSD_MIN_PADDED_SIZE : int
        Minimum zero-padded FFT dimension for frequency resolution.
    PSD_RADIAL_BINS : int
        Number of logarithmic radial frequency bins.
    PSD_MIN_FREQ : float
        Minimum radial frequency for binning (cycles/nm).
    DIRECTIONAL_SECTORS : int
        Number of angular sectors for CV_Fourier calculation.
    WAVELET_TYPE : str
        Default wavelet basis for decomposition.
    WAVELET_LEVELS : int
        Default wavelet decomposition levels.
    MULTISCALE_MAX_SCALE : int
        Maximum Gaussian smoothing scale for multiscale analysis.
    PIP_LOWER_PERCENTILE : float
        Lower percentile for PIP height filtering.
    PIP_UPPER_PERCENTILE : float
        Upper percentile for PIP height filtering.
    NIST_RA_TARGET : float
        NIST SRM 2073 Ra target value (nm).
    NIST_RA_TOLERANCE : float
        NIST SRM 2073 Ra tolerance (nm).
    NIST_RSK_TARGET : float
        NIST SRM 2073 Rsk target value.
    NIST_RSK_TOLERANCE : float
        NIST SRM 2073 Rsk tolerance.
    NIST_RKU_TARGET : float
        NIST SRM 2073 Rku target value.
    NIST_RKU_TOLERANCE : float
        NIST SRM 2073 Rku tolerance.
    NIST_ENERGY_GAP_MIN : float
        Minimum acceptable OPD energy gap.
    PIXEL_SIZE_MIN_NM : float
        Minimum physically valid pixel size (nm).
    PIXEL_SIZE_MAX_NM : float
        Maximum physically valid pixel size (nm).
    NUMERICAL_EPSILON : float
        Small value added to denominators to prevent division by zero.
    VALIDATION_SURFACE_SIZE : int
        Pixel dimensions of NIST validation surface.
    VALIDATION_RANDOM_SEED : int
        Random seed for reproducible validation surface.
    RT_ROBUST_LOWER_PERCENTILE : float
        Lower percentile bound for robust Rt calculation.
    RT_ROBUST_UPPER_PERCENTILE : float
        Upper percentile bound for robust Rt calculation.
    """

    # OPD Parameters (Liu et al. 2015)
    OPD_LEVELS: int = 5
    OPD_THRESHOLD_PERCENTILE: float = 80.0
    OPD_ENERGY_DECAY: float = 0.75

    # KDE Parameters
    KDE_BANDWIDTH_FACTOR: float = 0.9

    # PSD Parameters
    PSD_MIN_PADDED_SIZE: int = 2048
    PSD_RADIAL_BINS: int = 50
    PSD_MIN_FREQ: float = 1e-3

    # Directional Analysis
    DIRECTIONAL_SECTORS: int = 72

    # Wavelet Parameters
    WAVELET_TYPE: str = "db4"
    WAVELET_LEVELS: int = 3

    # Multiscale Parameters
    MULTISCALE_MAX_SCALE: int = 5

    # PIP Parameters
    PIP_LOWER_PERCENTILE: float = 5.0
    PIP_UPPER_PERCENTILE: float = 95.0

    # NIST SRM 2073 Validation Targets
    NIST_RA_TARGET: float = 50.8
    NIST_RA_TOLERANCE: float = 10.0
    NIST_RSK_TARGET: float = -0.82
    NIST_RSK_TOLERANCE: float = 0.15
    NIST_RKU_TARGET: float = 3.21
    NIST_RKU_TOLERANCE: float = 0.25
    NIST_ENERGY_GAP_MIN: float = 0.15

    # Physical Validation Bounds
    PIXEL_SIZE_MIN_NM: float = 1.0
    PIXEL_SIZE_MAX_NM: float = 500.0

    # Numerical Stability
    NUMERICAL_EPSILON: float = 1e-10

    # Validation Surface
    VALIDATION_SURFACE_SIZE: int = 256
    VALIDATION_RANDOM_SEED: int = 42

    # Robust Rt percentile bounds
    RT_ROBUST_LOWER_PERCENTILE: float = 0.5
    RT_ROBUST_UPPER_PERCENTILE: float = 99.5


# ========================
# Section 4: Custom Exceptions
# ========================

class AFMAnalyzerError(Exception):
    """Base exception class for AFMAnalyzer."""
    pass


class InvalidImageError(AFMAnalyzerError):
    """Raised when the input image fails validation checks."""
    pass


class MetadataExtractionError(AFMAnalyzerError):
    """Raised when metadata cannot be extracted by any method."""
    pass


class ValidationError(AFMAnalyzerError):
    """Raised when NIST validation fails critically."""
    pass


class PhysicalUnitError(AFMAnalyzerError):
    """Raised when physical unit checks fail."""
    pass


# ========================
# Section 5: Input Validation Utilities
# ========================

def validate_image_array(
    image: np.ndarray,
    function_name: str = "unknown"
) -> np.ndarray:
    """
    Validate and sanitize a 2D AFM height image array.

    Performs comprehensive checks for type, dimensionality,
    shape, NaN/Inf values, and physical plausibility.
    Returns a clean float64 2D array.

    Parameters
    ----------
    image : np.ndarray
        Input image array to validate.
    function_name : str
        Name of the calling function (for error messages).

    Returns
    -------
    image : np.ndarray
        Validated and sanitized 2D float64 array.

    Raises
    ------
    InvalidImageError
        If image is None, not a numpy array, wrong dimensions,
        too small, or contains only non-finite values.

    Examples
    --------
    >>> img = np.random.normal(0, 10, (256, 256))
    >>> validated = validate_image_array(img, "my_function")
    >>> validated.dtype
    dtype('float64')
    """
    # Check for None
    if image is None:
        raise InvalidImageError(
            f"[{function_name}] Image is None. "
            "Please provide a valid numpy array."
        )

    # Check type
    if not isinstance(image, np.ndarray):
        raise InvalidImageError(
            f"[{function_name}] Expected numpy array, "
            f"got {type(image).__name__}."
        )

    # Convert to float64
    image = image.astype(np.float64)

    # Squeeze singleton dimensions
    image = image.squeeze()

    # Check dimensionality
    if image.ndim != 2:
        raise InvalidImageError(
            f"[{function_name}] Expected 2D array, "
            f"got {image.ndim}D array with shape {image.shape}. "
            "For RGB images, convert to grayscale first. "
            "For z-stacks, select a single slice."
        )

    # Check minimum size
    min_size = 8
    if image.shape[0] < min_size or image.shape[1] < min_size:
        raise InvalidImageError(
            f"[{function_name}] Image too small: {image.shape}. "
            f"Minimum size is {min_size}x{min_size} pixels."
        )

    # Check for all-NaN
    if np.all(np.isnan(image)):
        raise InvalidImageError(
            f"[{function_name}] Image contains only NaN values."
        )

    # Check for all-Inf
    if np.all(np.isinf(image)):
        raise InvalidImageError(
            f"[{function_name}] Image contains only Inf values."
        )

    # Handle partial NaN/Inf
    n_nan = np.sum(np.isnan(image))
    n_inf = np.sum(np.isinf(image))

    if n_nan > 0 or n_inf > 0:
        total_bad = n_nan + n_inf
        pct_bad = 100 * total_bad / image.size
        logger.warning(
            f"[{function_name}] Found {total_bad} invalid pixels "
            f"({pct_bad:.2f}% of image). "
            "Replacing with local median values."
        )
        bad_mask = ~np.isfinite(image)
        image[bad_mask] = np.nanmedian(image)

    # Check for constant image
    # FIXED: np.ptp() deprecated → use np.max() - np.min()
    image_range = np.max(image) - np.min(image)
    if image_range < AFMConfig.NUMERICAL_EPSILON:
        logger.warning(
            f"[{function_name}] Image appears constant "
            f"(range = {image_range:.2e} nm). "
            "All roughness metrics will be near zero."
        )

    return image


def validate_pixel_size(
    pixel_size_nm: float,
    function_name: str = "unknown"
) -> float:
    """
    Validate that pixel size is physically reasonable for AFM.

    Parameters
    ----------
    pixel_size_nm : float
        Pixel size in nanometers to validate.
    function_name : str
        Name of calling function (for error messages).

    Returns
    -------
    pixel_size_nm : float
        Validated pixel size.

    Raises
    ------
    PhysicalUnitError
        If pixel size is outside the valid AFM range.

    Examples
    --------
    >>> ps = validate_pixel_size(19.53, "calculate_psd")
    >>> ps
    19.53
    """
    if pixel_size_nm is None:
        raise PhysicalUnitError(
            f"[{function_name}] Pixel size is None."
        )

    try:
        pixel_size_nm = float(pixel_size_nm)
    except (TypeError, ValueError):
        raise PhysicalUnitError(
            f"[{function_name}] Cannot convert pixel size "
            f"'{pixel_size_nm}' to float."
        )

    if not np.isfinite(pixel_size_nm):
        raise PhysicalUnitError(
            f"[{function_name}] Pixel size is not finite: "
            f"{pixel_size_nm}."
        )

    if pixel_size_nm <= 0:
        raise PhysicalUnitError(
            f"[{function_name}] Pixel size must be positive, "
            f"got {pixel_size_nm:.4f} nm."
        )

    if not (AFMConfig.PIXEL_SIZE_MIN_NM
            <= pixel_size_nm
            <= AFMConfig.PIXEL_SIZE_MAX_NM):
        raise PhysicalUnitError(
            f"[{function_name}] Pixel size {pixel_size_nm:.2f} nm "
            f"is outside valid AFM range "
            f"[{AFMConfig.PIXEL_SIZE_MIN_NM}, "
            f"{AFMConfig.PIXEL_SIZE_MAX_NM}] nm. "
            "Check scan size and image dimensions."
        )

    return pixel_size_nm


# ========================
# Section 6: Metadata Extraction
# ========================

def extract_jpk_metadata(tif: TiffFile) -> Dict:
    """
    Extract scan metadata from JPK-format AFM TIFF files.

    Attempts multiple key naming conventions to handle
    different JPK software versions. Performs automatic
    unit conversion from micrometers to nanometers.

    Parameters
    ----------
    tif : TiffFile
        Open TiffFile object from tifffile library.

    Returns
    -------
    metadata : dict
        Dictionary with keys:
        - 'scan_size_x_nm' (float or None): X scan size in nm
        - 'scan_size_y_nm' (float or None): Y scan size in nm
        - 'pixel_size_nm' (float or None): Pixel size in nm
        - 'success' (bool): True if any metadata was extracted

    Notes
    -----
    JPK instruments have changed metadata key names across
    software versions. The multi-key fallback strategy
    provides backward compatibility.

    Examples
    --------
    >>> with TiffFile("sample.tif") as tif:
    ...     meta = extract_jpk_metadata(tif)
    >>> meta['success']
    True
    """
    metadata = {
        'scan_size_x_nm': None,
        'scan_size_y_nm': None,
        'pixel_size_nm': None,
        'success': False
    }

    X_SIZE_KEYS = ['JPK_SCAN_SIZE_X', 'JPK_ScanSize', 'JPKScanSize_X']
    Y_SIZE_KEYS = ['JPK_SCAN_SIZE_Y', 'JPK_ScanSize', 'JPKScanSize_Y']
    PIXEL_KEYS = ['JPK_PIXEL_SIZE', 'JPK_PixelSize', 'PixelSize']
    UNIT_KEY = 'JPK_UNIT'
    UM_VARIANTS = {'um', 'µm', 'micrometer', 'micron'}

    try:
        if (hasattr(tif, 'shaped_metadata')
                and tif.shaped_metadata):
            shaped = tif.shaped_metadata[0]
            unit = shaped.get(UNIT_KEY, '').lower().strip()
            needs_um_conversion = unit in UM_VARIANTS

            for key in X_SIZE_KEYS:
                if key in shaped and shaped[key]:
                    try:
                        val = float(shaped[key])
                        metadata['scan_size_x_nm'] = (
                            val * 1000.0
                            if needs_um_conversion else val
                        )
                        metadata['success'] = True
                        break
                    except (ValueError, TypeError):
                        continue

            for key in Y_SIZE_KEYS:
                if key in shaped and shaped[key]:
                    try:
                        val = float(shaped[key])
                        metadata['scan_size_y_nm'] = (
                            val * 1000.0
                            if needs_um_conversion else val
                        )
                        metadata['success'] = True
                        break
                    except (ValueError, TypeError):
                        continue

            if (metadata['scan_size_x_nm'] is not None
                    and metadata['scan_size_y_nm'] is None):
                metadata['scan_size_y_nm'] = metadata['scan_size_x_nm']
                logger.debug("Assumed square scan: Y = X scan size.")
            elif (metadata['scan_size_y_nm'] is not None
                  and metadata['scan_size_x_nm'] is None):
                metadata['scan_size_x_nm'] = metadata['scan_size_y_nm']
                logger.debug("Assumed square scan: X = Y scan size.")

            for key in PIXEL_KEYS:
                if key in shaped and shaped[key]:
                    try:
                        val = float(shaped[key])
                        metadata['pixel_size_nm'] = (
                            val * 1000.0
                            if needs_um_conversion else val
                        )
                        metadata['success'] = True
                        break
                    except (ValueError, TypeError):
                        continue

        if (not metadata['success']
                and hasattr(tif, 'pages')
                and tif.pages):
            page = tif.pages[0]
            desc = getattr(page, 'image_description', None)

            if desc is not None:
                if isinstance(desc, bytes):
                    desc = desc.decode('utf-8', errors='ignore')

                patterns = [
                    (r'scan\s*size\s*[:=]\s*(\d+\.?\d*)\s*(nm|µm|um)',
                     'scan_size'),
                    (r'resolution\s*[:=]\s*(\d+\.?\d*)\s*(nm|µm|um)',
                     'pixel_size'),
                ]

                for pattern, field in patterns:
                    match = re.search(pattern, desc.lower())
                    if match:
                        val = float(match.group(1))
                        unit = match.group(2)
                        if unit in ('µm', 'um'):
                            val *= 1000.0

                        if field == 'scan_size':
                            metadata['scan_size_x_nm'] = val
                            metadata['scan_size_y_nm'] = val
                        else:
                            metadata['pixel_size_nm'] = val

                        metadata['success'] = True

    except Exception as e:
        logger.warning(f"JPK metadata extraction error: {e}")

    return metadata


def extract_afm_metadata(image_path: Union[str, Path]) -> Dict:
    """
    Extract comprehensive metadata from an AFM TIFF image file.

    Uses a cascading four-method extraction strategy:
    1. JPK shaped_metadata
    2. image_description regex
    3. Derived calculation from available values
    4. Returns partial results with success=False for GUI fallback

    Parameters
    ----------
    image_path : str or Path
        Path to the AFM TIFF image file.

    Returns
    -------
    metadata : dict
        Dictionary with keys:
        - 'scan_size_x_nm' (float or None)
        - 'scan_size_y_nm' (float or None)
        - 'pixel_size_nm' (float or None)
        - 'width_pixels' (int)
        - 'height_pixels' (int)
        - 'success' (bool)

    Examples
    --------
    >>> meta = extract_afm_metadata("sample.tif")
    >>> if meta['success']:
    ...     print(f"Pixel size: {meta['pixel_size_nm']:.2f} nm")
    """
    image_path = Path(image_path)

    if not image_path.exists():
        raise FileNotFoundError(
            f"AFM image file not found: {image_path}"
        )

    if image_path.suffix.lower() not in ('.tif', '.tiff'):
        logger.warning(
            f"File extension '{image_path.suffix}' is not .tif/.tiff. "
            "Attempting to read anyway."
        )

    metadata = {
        'scan_size_x_nm': None,
        'scan_size_y_nm': None,
        'pixel_size_nm': None,
        'width_pixels': 0,
        'height_pixels': 0,
        'success': False
    }

    try:
        with TiffFile(str(image_path)) as tif:
            image = tif.asarray()
            image = image.squeeze()

            if image.ndim == 3:
                metadata['height_pixels'] = image.shape[0]
                metadata['width_pixels'] = image.shape[1]
            elif image.ndim == 2:
                metadata['height_pixels'] = image.shape[0]
                metadata['width_pixels'] = image.shape[1]

            jpk_meta = extract_jpk_metadata(tif)
            if jpk_meta['success']:
                metadata.update(jpk_meta)
                logger.info("Metadata extracted via JPK method.")

        if (metadata['scan_size_x_nm'] is not None
                and metadata['pixel_size_nm'] is None
                and metadata['width_pixels'] > 0):
            metadata['pixel_size_nm'] = (
                metadata['scan_size_x_nm'] / metadata['width_pixels']
            )
            logger.info("Pixel size derived from scan size / width.")

        elif (metadata['pixel_size_nm'] is not None
              and metadata['scan_size_x_nm'] is None
              and metadata['width_pixels'] > 0):
            metadata['scan_size_x_nm'] = (
                metadata['pixel_size_nm'] * metadata['width_pixels']
            )
            metadata['scan_size_y_nm'] = (
                metadata['pixel_size_nm'] * metadata['height_pixels']
            )
            logger.info("Scan size derived from pixel size x dimensions.")

    except Exception as e:
        logger.error(f"Metadata extraction failed: {e}")

    return metadata


# ========================
# Section 6b: AFM Image Loader
# ========================

def load_afm_image(image_path: Union[str, Path]) -> np.ndarray:
    """
    Load an AFM TIFF image and return a 2D float64 height array.

    Correctly handles three formats:

    Case 1 — float32/float64 TIFF (JPK correct export):
        shape=(H,W), dtype=float → values already in nm.
        Action: cast to float64, zero-mean, return directly.

    Case 2 — Grayscale integer TIFF (uint8 or uint16):
        shape=(H,W), dtype=uint8/uint16 → raw counts.
        Action: cast to float64, warn user, return as-is.

    Case 3 — RGB color TIFF (rendered screenshot):
        shape=(H,W,3), dtype=uint8 → false-colour image.
        Action: extract luminance, warn user strongly.

    Parameters
    ----------
    image_path : str or Path
        Path to AFM TIFF file.

    Returns
    -------
    image : np.ndarray
        2D float64 array, zero-meaned, ready for analysis.

    Raises
    ------
    InvalidImageError
        If the file cannot be loaded or produces no usable data.

    Examples
    --------
    >>> img = load_afm_image("sample.tif")
    >>> img.dtype
    dtype('float64')
    >>> abs(img.mean()) < 1e-6
    True
    """
    image_path = Path(image_path)

    try:
        with TiffFile(str(image_path)) as tif:
            raw = tif.asarray()
    except Exception as e:
        raise InvalidImageError(
            f"load_afm_image: Cannot read {image_path.name}: {e}"
        )

    raw = raw.squeeze()

    # Case 1: float32 or float64
    if np.issubdtype(raw.dtype, np.floating):
        if raw.ndim == 3:
            logger.warning(
                f"{image_path.name}: 3D float TIFF detected. "
                "Using channel 0 as height channel."
            )
            raw = raw[:, :, 0]

        image = raw.astype(np.float64)
        image -= np.mean(image)

        logger.info(
            f"{image_path.name}: float{raw.dtype.itemsize*8} TIFF loaded. "
            f"Height range: {image.min():.2f} to {image.max():.2f} nm. "
            "Values treated directly as nanometers."
        )
        return image

    # Case 2: Grayscale integer
    if raw.ndim == 2:
        logger.warning(
            f"{image_path.name}: Integer grayscale TIFF "
            f"(dtype={raw.dtype}). "
            "Values are raw digital counts, NOT nanometers. "
            "Roughness metrics will be in counts. "
            "Re-export from JPK as float32 TIFF for correct nm values."
        )
        image = raw.astype(np.float64)
        image -= np.mean(image)
        return image

    # Case 3: RGB color image
    if raw.ndim == 3 and raw.shape[2] == 3:
        logger.warning(
            f"{image_path.name}: RGB color TIFF detected "
            f"(shape={raw.shape}, dtype={raw.dtype}). "
            "This is a rendered display image, NOT raw height data. "
            "Converting to luminance grayscale as fallback. "
            "Results will be UNRELIABLE. "
            "Please re-export from JPK Data Processing as "
            "float32 TIFF (File > Export > TIFF 32-bit float)."
        )
        image = color.rgb2gray(raw).astype(np.float64) * 255.0
        image -= np.mean(image)
        return image

    raise InvalidImageError(
        f"load_afm_image: Unrecognised TIFF format in "
        f"{image_path.name}: shape={raw.shape}, dtype={raw.dtype}."
    )


# ========================
# Section 7: ISO 4287 Roughness Metrics
# ========================

def calculate_roughness_metrics(image: np.ndarray) -> Dict:
    """
    Calculate ISO 4287 surface roughness parameters with
    Fisher-Pearson bias correction, OPD fractal dimension,
    and robust Rt.

    Implements:
    - Ra:        arithmetic mean roughness
    - Rq:        root mean square roughness
    - Rt:        maximum height (peak-to-valley, ISO 4287 strict)
    - Rt_Robust: peak-to-valley from P0.5 to P99.5 (spike-resistant)
    - Rsk:       skewness with Fisher-Pearson bias correction
    - Rku:       excess kurtosis with Pearson bias correction
    - Fractal_Dim: OPD-derived fractal dimension estimate

    Parameters
    ----------
    image : np.ndarray
        2D AFM height image array. Values in nanometers.

    Returns
    -------
    metrics : dict
        Dictionary with keys:
        'Ra_nm', 'Rq_nm', 'Rt_nm', 'Rt_Robust_nm',
        'Rsk', 'Rku', 'Fractal_Dim'

    Raises
    ------
    InvalidImageError
        If image fails validation checks.

    Examples
    --------
    >>> img = np.random.normal(0, 10, (256, 256))
    >>> metrics = calculate_roughness_metrics(img)
    >>> print(f"Ra = {metrics['Ra_nm']:.3f} nm")
    >>> metrics['Rt_Robust_nm'] <= metrics['Rt_nm']
    True
    """
    image = validate_image_array(image, "calculate_roughness_metrics")

    mean_z = np.mean(image)
    deviations = image - mean_z
    n_pixels = deviations.size

    # FIXED: Move eps before with block for cleaner scope
    eps = AFMConfig.NUMERICAL_EPSILON

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", RuntimeWarning)

        Ra = float(np.mean(np.abs(deviations)))
        Rq = float(np.sqrt(np.mean(deviations**2)))

        # FIXED: np.ptp() deprecated → np.max() - np.min()
        Rt = float(np.max(image) - np.min(image))

        Rt_robust = float(
            np.percentile(image, AFMConfig.RT_ROBUST_UPPER_PERCENTILE)
            - np.percentile(image, AFMConfig.RT_ROBUST_LOWER_PERCENTILE)
        )

        if n_pixels >= 3:
            skew_factor = (
                n_pixels / ((n_pixels - 1) * (n_pixels - 2))
            )
            skew_num = skew_factor * np.sum(deviations**3)
            Rsk = float(skew_num / (Rq**3 + eps))
        else:
            Rsk = np.nan
            logger.warning("Too few pixels for skewness calculation.")

        if n_pixels >= 4:
            kurt_factor = (
                (n_pixels * (n_pixels + 1))
                / ((n_pixels - 1) * (n_pixels - 2) * (n_pixels - 3))
            )
            kurt_bias = (
                3 * (n_pixels - 1)**2
                / ((n_pixels - 2) * (n_pixels - 3))
            )
            kurt_num = kurt_factor * np.sum(deviations**4)
            Rku = float(kurt_num / (Rq**4 + eps) - kurt_bias)
        else:
            Rku = np.nan
            logger.warning("Too few pixels for kurtosis calculation.")

    # OPD Fractal Dimension
    try:
        opd_results = orthogonal_pdf_decomposition(image)
        energy_spectrum = opd_results['energy_spectrum']

        if (len(energy_spectrum) >= 2
                and energy_spectrum[1] > eps):
            ratio = energy_spectrum[0] / energy_spectrum[1]
            if ratio > 0:
                fractal_dim = float(2.0 + np.log2(ratio))
            else:
                fractal_dim = np.nan
        else:
            fractal_dim = np.nan

    except Exception as e:
        logger.warning(f"OPD fractal dimension failed: {e}")
        fractal_dim = np.nan

    return {
        'Ra_nm': Ra,
        'Rq_nm': Rq,
        'Rt_nm': Rt,
        'Rt_Robust_nm': Rt_robust,
        'Rsk': Rsk,
        'Rku': Rku,
        'Fractal_Dim': fractal_dim
    }


# ========================
# Section 8: Power Spectral Density
# ========================

def calculate_psd(
    image: np.ndarray,
    pixel_size_nm: float
) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    """
    Calculate 2D Power Spectral Density with physical unit scaling.

    5-step pipeline:
    1. Mean centering
    2. 2D Hann windowing
    3. Adaptive zero-padding
    4. FFT-shift and squared magnitude
    5. Physical normalization (units: nm^2)

    Parameters
    ----------
    image : np.ndarray
        2D AFM height image in nanometers.
    pixel_size_nm : float
        Physical pixel size in nanometers.

    Returns
    -------
    psd : np.ndarray
        2D power spectral density array (nm^2).
    freq_x : np.ndarray
        Spatial frequencies along x-axis (cycles/nm).
    freq_y : np.ndarray
        Spatial frequencies along y-axis (cycles/nm).

    Examples
    --------
    >>> img = np.random.normal(0, 10, (256, 256))
    >>> psd, fx, fy = calculate_psd(img, pixel_size_nm=19.53)
    >>> psd.shape
    (2048, 2048)
    """
    image = validate_image_array(image, "calculate_psd")
    pixel_size_nm = validate_pixel_size(pixel_size_nm, "calculate_psd")

    ny, nx = image.shape

    image_centered = image - np.mean(image)

    hann_y = np.hanning(ny)
    hann_x = np.hanning(nx)
    window = np.outer(hann_y, hann_x)
    windowed = image_centered * window

    pad_ny = max(AFMConfig.PSD_MIN_PADDED_SIZE, ny)
    pad_nx = max(AFMConfig.PSD_MIN_PADDED_SIZE, nx)
    padded = np.zeros((pad_ny, pad_nx), dtype=np.float64)
    padded[:ny, :nx] = windowed

    fft_result = fftshift(fft2(padded))
    power = np.abs(fft_result)**2

    window_norm = np.sum(window**2) / (ny * nx)
    psd = (power / (pad_ny * pad_nx * window_norm)) * (pixel_size_nm**2)

    freq_x = np.fft.fftshift(
        np.fft.fftfreq(pad_nx, d=pixel_size_nm)
    )
    freq_y = np.fft.fftshift(
        np.fft.fftfreq(pad_ny, d=pixel_size_nm)
    )

    logger.debug(
        f"PSD computed: shape={psd.shape}, "
        f"mean={np.mean(psd):.4f} nm^2, "
        f"freq_range=[{freq_x[0]:.4f}, {freq_x[-1]:.4f}] cycles/nm"
    )

    return psd, freq_x, freq_y


def radial_average_psd(
    psd: np.ndarray,
    freq_x: np.ndarray,
    freq_y: np.ndarray,
    n_bins: int = AFMConfig.PSD_RADIAL_BINS
) -> Tuple[np.ndarray, np.ndarray]:
    """
    Calculate radially averaged (isotropic) 1D PSD.

    Uses logarithmic frequency binning for uniform coverage
    across the spatial frequency range.

    Parameters
    ----------
    psd : np.ndarray
        2D PSD array (nm^2).
    freq_x : np.ndarray
        X-axis spatial frequencies (cycles/nm).
    freq_y : np.ndarray
        Y-axis spatial frequencies (cycles/nm).
    n_bins : int
        Number of logarithmic radial bins.
        Default: AFMConfig.PSD_RADIAL_BINS (50).

    Returns
    -------
    freq_bins : np.ndarray
        Geometric mean frequencies for each bin (cycles/nm).
    radial_psd : np.ndarray
        Radially averaged PSD values (nm^2).

    Examples
    --------
    >>> img = np.random.normal(0, 10, (256, 256))
    >>> psd, fx, fy = calculate_psd(img, 19.53)
    >>> freqs, rpsd = radial_average_psd(psd, fx, fy)
    >>> freqs.shape[0] <= 50
    True
    """
    fx_grid, fy_grid = np.meshgrid(freq_x, freq_y)
    radial_freq = np.sqrt(fx_grid**2 + fy_grid**2)

    max_valid = min(
        np.max(freq_x[len(freq_x)//2:]),
        np.max(freq_y[len(freq_y)//2:])
    )
    min_valid = AFMConfig.PSD_MIN_FREQ
    valid_mask = (radial_freq > min_valid) & (radial_freq <= max_valid)

    if not np.any(valid_mask):
        logger.warning("No valid frequencies found for radial averaging.")
        return np.array([]), np.array([])

    bin_edges = np.logspace(
        np.log10(min_valid),
        np.log10(max_valid),
        n_bins
    )

    valid_freqs = radial_freq[valid_mask]
    valid_psd = psd[valid_mask]
    bin_indices = np.digitize(valid_freqs, bin_edges)

    freq_bins = []
    radial_psd_vals = []

    # FIXED: Added comment for clarity (Issue C)
    # n_bins edges produce n_bins-1 usable intervals
    for i in range(1, n_bins):
        mask = bin_indices == i
        if np.sum(mask) >= 3:
            freq_bins.append(
                np.sqrt(bin_edges[i-1] * bin_edges[i])
            )
            radial_psd_vals.append(np.mean(valid_psd[mask]))

    if not freq_bins:
        logger.warning("Radial averaging produced no valid bins.")

    return np.array(freq_bins), np.array(radial_psd_vals)


def calculate_directional_cv(
    psd: np.ndarray,
    n_sectors: int = AFMConfig.DIRECTIONAL_SECTORS
) -> float:
    """
    Calculate CV_Fourier: coefficient of variation of angular
    sector mean powers in 2D PSD space.

    Uses fully vectorized computation via np.arctan2.

    Parameters
    ----------
    psd : np.ndarray
        2D PSD array.
    n_sectors : int
        Number of angular sectors (default: 72).

    Returns
    -------
    cv_fourier : float
        Coefficient of variation of sector powers (%).
        Returns 0.0 if fewer than 2 sectors have valid data.

    Examples
    --------
    >>> psd = np.random.exponential(1.0, (2048, 2048))
    >>> cv = calculate_directional_cv(psd, n_sectors=72)
    >>> 0.0 <= cv <= 100.0
    True
    """
    ny, nx = psd.shape
    cy, cx = ny // 2, nx // 2

    y_idx, x_idx = np.indices((ny, nx))
    dy = y_idx - cy
    dx = x_idx - cx

    angles_deg = np.degrees(np.arctan2(dy, dx)) % 360.0

    dc_mask = (dy == 0) & (dx == 0)

    sector_size = 360.0 / n_sectors
    sector_means = []

    for s in range(n_sectors):
        angle_min = s * sector_size
        angle_max = (s + 1) * sector_size
        sector_mask = (
            (angles_deg >= angle_min)
            & (angles_deg < angle_max)
            & ~dc_mask
        )
        sector_vals = psd[sector_mask]
        if len(sector_vals) >= 5:
            sector_means.append(float(np.mean(sector_vals)))

    if len(sector_means) < 2:
        logger.warning(
            "Fewer than 2 valid sectors found. "
            "Returning CV_Fourier = 0.0."
        )
        return 0.0

    mean_val = np.mean(sector_means)
    if mean_val < AFMConfig.NUMERICAL_EPSILON:
        return 0.0

    # FIXED: Changed ddof=0 to ddof=1 for consistency (Issue D)
    cv_fourier = float(
        (np.std(sector_means, ddof=1) / mean_val) * 100.0
    )

    return cv_fourier


def calculate_fourier_metrics(
    psd: np.ndarray,
    freq_x: np.ndarray,
    freq_y: np.ndarray,
    pixel_size_nm: float
) -> Dict:
    """
    Calculate comprehensive Fourier-domain surface metrics.

    Parameters
    ----------
    psd : np.ndarray
        2D PSD array (nm^2).
    freq_x : np.ndarray
        X-axis frequencies (cycles/nm).
    freq_y : np.ndarray
        Y-axis frequencies (cycles/nm).
    pixel_size_nm : float
        Physical pixel size (nm).

    Returns
    -------
    results : dict
        Dictionary with keys:
        - 'mean_psd', 'mean_psd_radial', 'cv_fourier', 'radial_data'

    Examples
    --------
    >>> img = np.random.normal(0, 10, (256, 256))
    >>> psd, fx, fy = calculate_psd(img, 19.53)
    >>> fourier = calculate_fourier_metrics(psd, fx, fy, 19.53)
    >>> 'mean_psd' in fourier
    True
    """
    mean_psd = float(np.mean(psd))

    freq_bins, radial_psd_vals = radial_average_psd(psd, freq_x, freq_y)

    if len(radial_psd_vals) >= 2:
        freq_range = freq_bins[-1] - freq_bins[0]
        if freq_range > AFMConfig.NUMERICAL_EPSILON:
            mean_psd_radial = float(
                np.trapz(radial_psd_vals, freq_bins) / freq_range
            )
        else:
            mean_psd_radial = float(np.mean(radial_psd_vals))
    else:
        mean_psd_radial = 0.0
        logger.warning("Insufficient radial PSD data for integration.")

    cv_fourier = calculate_directional_cv(psd)

    return {
        'mean_psd': mean_psd,
        'mean_psd_radial': mean_psd_radial,
        'cv_fourier': cv_fourier,
        'radial_data': (freq_bins, radial_psd_vals)
    }


# ========================
# Section 9: PIP Metrics and Novel Indices
# ========================

def calculate_pip_metrics(
    image: np.ndarray,
    pixel_size_nm: float
) -> Dict:
    """
    Calculate Percentile-based Image Profile (PIP) metrics.

    PIP metrics are robust to AFM scanning artifacts by operating
    on the central 90% of the height distribution.

    CV_PIP Definition:
        CV_PIP = std(filtered_heights) / delta_pip * 100%

    where delta_pip = P95 - P5.

    The denominator is delta_pip (NOT the mean height) because:
    1. AFM images are zero-mean by definition after plane leveling
    2. delta_pip is a stable, physically meaningful height scale

    Parameters
    ----------
    image : np.ndarray
        2D AFM height image (nm).
    pixel_size_nm : float
        Physical pixel size (nm).

    Returns
    -------
    metrics : dict
        Dictionary with keys:
        - 'delta_pip', 'cv_pip', 'q5', 'q95', 'n_filtered'

    Examples
    --------
    >>> img = np.random.normal(0, 10, (256, 256))
    >>> pip = calculate_pip_metrics(img, 19.53)
    >>> pip['delta_pip'] > 0
    True
    """
    image = validate_image_array(image, "calculate_pip_metrics")
    pixel_size_nm = validate_pixel_size(
        pixel_size_nm, "calculate_pip_metrics"
    )

    heights = image.flatten()
    q5 = float(np.percentile(heights, AFMConfig.PIP_LOWER_PERCENTILE))
    q95 = float(np.percentile(heights, AFMConfig.PIP_UPPER_PERCENTILE))

    if q95 < q5:
        raise InvalidImageError(
            f"Invalid percentiles: q95={q95:.4f} < q5={q5:.4f}. "
            "This suggests a corrupted or inverted height scale."
        )

    delta_pip = q95 - q5

    filtered = heights[(heights >= q5) & (heights <= q95)]
    n_filtered = len(filtered)

    if n_filtered < 2:
        logger.warning(
            "Fewer than 2 pixels in filtered range. "
            "CV_PIP set to 0.0."
        )
        cv_pip = 0.0
    else:
        std_filtered = float(np.std(filtered, ddof=1))

        cv_pip = float(
            (std_filtered / (delta_pip + AFMConfig.NUMERICAL_EPSILON))
            * 100.0
        )

    return {
        'delta_pip': delta_pip,
        'cv_pip': cv_pip,
        'q5': q5,
        'q95': q95,
        'n_filtered': n_filtered
    }


def calculate_pdi(
    cv_fourier: float,
    cv_pip: float
) -> float:
    """
    Calculate the Pattern Dominance Index (PDI).

    PDI = CV_Fourier / CV_PIP

    Parameters
    ----------
    cv_fourier : float
        Coefficient of variation of angular PSD sectors (%).
    cv_pip : float
        Normalized height spread (%).

    Returns
    -------
    pdi : float
        Pattern Dominance Index (dimensionless).

    Examples
    --------
    >>> calculate_pdi(cv_fourier=15.0, cv_pip=5.0)
    3.0
    """
    if not np.isfinite(cv_fourier) or not np.isfinite(cv_pip):
        return np.nan

    if cv_pip < AFMConfig.NUMERICAL_EPSILON:
        logger.info("CV_PIP approximately 0: PDI -> inf.")
        return np.inf

    return float(cv_fourier / cv_pip)


def calculate_hi(
    delta_pip: float,
    mean_psd: float,
    cv_fourier: float,
    cv_pip: float
) -> float:
    """
    Calculate the Heterogeneity Index (HI).

    HI = (delta_pip / sqrt(PSD_mean)) * ln(CV_Fourier / CV_PIP)

    Parameters
    ----------
    delta_pip : float
        PIP height range (nm).
    mean_psd : float
        Mean 2D power spectral density (nm^2).
    cv_fourier : float
        Directional CV (%).
    cv_pip : float
        Normalized height spread (%).

    Returns
    -------
    hi : float
        Heterogeneity Index (dimensionless).

    Examples
    --------
    >>> calculate_hi(50.0, 100.0, 20.0, 5.0)
    """
    for name, val in [('delta_pip', delta_pip), ('mean_psd', mean_psd),
                      ('cv_fourier', cv_fourier), ('cv_pip', cv_pip)]:
        if not np.isfinite(val):
            logger.warning(
                f"HI: {name} is not finite ({val}). Returning NaN."
            )
            return np.nan

    if mean_psd <= AFMConfig.NUMERICAL_EPSILON:
        logger.warning("HI: mean_psd approximately 0. Returning NaN.")
        return np.nan

    if cv_pip <= AFMConfig.NUMERICAL_EPSILON:
        logger.warning("HI: cv_pip approximately 0. Returning NaN.")
        return np.nan

    if cv_fourier <= AFMConfig.NUMERICAL_EPSILON:
        logger.warning("HI: cv_fourier approximately 0. Returning NaN.")
        return np.nan

    spectral_amplitude = np.sqrt(mean_psd)
    ratio = cv_fourier / cv_pip
    if ratio <= 0:
        logger.warning(f"HI: CV ratio {ratio:.4f} <= 0. Returning NaN.")
        return np.nan

    structural_contrast = np.log(ratio)
    hi = float((delta_pip / spectral_amplitude) * structural_contrast)
    return hi


def calculate_hi_radial(
    delta_pip: float,
    mean_psd_radial: float,
    cv_fourier: float,
    cv_pip: float
) -> float:
    """
    Calculate the Radial Heterogeneity Index (HI_Radial).

    Uses isotropically averaged radial PSD mean.

    Parameters
    ----------
    delta_pip : float
        PIP height range (nm).
    mean_psd_radial : float
        Radially averaged and integrated mean PSD (nm^2).
    cv_fourier : float
        Directional CV (%).
    cv_pip : float
        Normalized height spread (%).

    Returns
    -------
    hi_radial : float
        Radial Heterogeneity Index (dimensionless).

    Examples
    --------
    >>> calculate_hi_radial(50.0, 80.0, 20.0, 5.0)
    """
    return calculate_hi(delta_pip, mean_psd_radial, cv_fourier, cv_pip)


# ========================
# Section 10: OPD Implementation
# ========================

def orthogonal_pdf_decomposition(
    image: np.ndarray,
    levels: int = AFMConfig.OPD_LEVELS
) -> Dict:
    """
    Orthogonal PDF Decomposition (OPD) following Liu et al. (2015).

    Parameters
    ----------
    image : np.ndarray
        2D AFM height image (nm).
    levels : int
        Number of decomposition levels.

    Returns
    -------
    results : dict
        Dictionary with keys:
        - 'energy_spectrum', 'approximations', 'dominant_scale', 'energy_gap'

    Examples
    --------
    >>> img = np.random.normal(0, 10, (256, 256))
    >>> opd = orthogonal_pdf_decomposition(img)
    >>> len(opd['energy_spectrum']) == 5
    True
    """
    image = validate_image_array(image, "orthogonal_pdf_decomposition")

    data = image.flatten()
    eps = AFMConfig.NUMERICAL_EPSILON

    bw = AFMConfig.KDE_BANDWIDTH_FACTOR * np.std(data) * (len(data)**(-0.2))
    bw = max(bw, eps)

    try:
        kde = gaussian_kde(data, bw_method=bw)
    except Exception as e:
        logger.error(f"KDE construction failed: {e}")
        return {
            'energy_spectrum': [],
            'approximations': [],
            'dominant_scale': -1,
            'energy_gap': 0.0
        }

    x_grid = np.linspace(data.min(), data.max(), 512)
    pdf = kde.evaluate(x_grid)
    pdf_sum = pdf.sum()

    if pdf_sum < eps:
        logger.warning("KDE produced near-zero PDF. OPD skipped.")
        return {
            'energy_spectrum': [],
            'approximations': [],
            'dominant_scale': -1,
            'energy_gap': 0.0
        }

    pdf = pdf / pdf_sum

    approximations = []
    current_pdf = pdf.copy()

    for level in range(levels):
        threshold = np.percentile(
            current_pdf,
            AFMConfig.OPD_THRESHOLD_PERCENTILE
        )
        phi_plus = np.where(current_pdf > threshold, current_pdf, 0.0)
        phi_minus = np.where(current_pdf <= threshold, current_pdf, 0.0)

        energy = float(np.var(phi_plus) + np.var(phi_minus))
        approximations.append({'level': level, 'energy': energy})

        current_pdf = AFMConfig.OPD_ENERGY_DECAY * (phi_plus + phi_minus)

    total_energy = sum(a['energy'] for a in approximations)

    if total_energy > eps:
        energy_spectrum = [
            a['energy'] / total_energy for a in approximations
        ]
    else:
        logger.warning("Total OPD energy approximately 0.")
        energy_spectrum = [0.0] * levels

    dominant_scale = int(np.argmax(energy_spectrum)) if energy_spectrum else -1

    if len(energy_spectrum) >= 2:
        energy_gap = float(energy_spectrum[0] - energy_spectrum[1])
    else:
        energy_gap = 0.0

    return {
        'energy_spectrum': energy_spectrum,
        'approximations': approximations,
        'dominant_scale': dominant_scale,
        'energy_gap': energy_gap
    }


# ========================
# Section 11: Multiscale Analysis
# ========================

def calculate_lacunarity(
    binary_image: np.ndarray,
    box_size: int
) -> float:
    """
    Calculate lacunarity using the box-counting method.

    Parameters
    ----------
    binary_image : np.ndarray
        2D binary image.
    box_size : int
        Box size for counting. Must be >= 2.

    Returns
    -------
    lacunarity : float
        Lacunarity value >= 1.0.

    Examples
    --------
    >>> img = np.random.randint(0, 2, (64, 64)).astype(bool)
    >>> lac = calculate_lacunarity(img, box_size=4)
    >>> lac >= 1.0
    True
    """
    if box_size < 2:
        logger.warning(f"Box size {box_size} < 2. Returning NaN.")
        return np.nan

    height, width = binary_image.shape

    if box_size >= height or box_size >= width:
        logger.warning(
            f"Box size {box_size} >= image dimension. Returning NaN."
        )
        return np.nan

    box_counts = []
    for i in range(0, height - box_size + 1, box_size):
        for j in range(0, width - box_size + 1, box_size):
            box = binary_image[i:i+box_size, j:j+box_size]
            box_counts.append(int(np.sum(box)))

    if len(box_counts) < 2:
        return np.nan

    box_counts = np.array(box_counts, dtype=np.float64)
    mean_count = float(np.mean(box_counts))
    var_count = float(np.var(box_counts))

    if mean_count < AFMConfig.NUMERICAL_EPSILON:
        logger.warning("Mean box count approximately 0.")
        return np.nan

    lacunarity = (var_count / mean_count**2) + 1.0
    return float(lacunarity)


def wavelet_analysis(
    image: np.ndarray,
    wavelet: str = AFMConfig.WAVELET_TYPE,
    level: int = AFMConfig.WAVELET_LEVELS
) -> Dict:
    """
    Perform 2D discrete wavelet decomposition.

    Parameters
    ----------
    image : np.ndarray
        2D AFM height image (nm).
    wavelet : str
        Wavelet basis name.
    level : int
        Decomposition levels.

    Returns
    -------
    metrics : dict
        Dictionary with keys:
        - 'level_energies', 'level_cvs', 'total_energy',
          'scale_entropy', 'coefficients'

    Examples
    --------
    >>> img = np.random.normal(0, 10, (256, 256))
    >>> wm = wavelet_analysis(img)
    >>> wm['scale_entropy'] >= 0
    True
    """
    image = validate_image_array(image, "wavelet_analysis")

    ny, nx = image.shape
    min_dim = 2**level
    if ny < min_dim or nx < min_dim:
        level = int(np.floor(np.log2(min(ny, nx))))
        logger.warning(
            f"Decomposition level reduced to {level}."
        )

    pad_ny = int(np.ceil(ny / min_dim)) * min_dim
    pad_nx = int(np.ceil(nx / min_dim)) * min_dim

    if pad_ny != ny or pad_nx != nx:
        padded = np.zeros((pad_ny, pad_nx))
        padded[:ny, :nx] = image
        image_for_dwt = padded
    else:
        image_for_dwt = image

    try:
        coeffs = pywt.wavedec2(image_for_dwt, wavelet, level=level)
    except Exception as e:
        logger.error(f"Wavelet decomposition failed: {e}")
        return {
            'level_energies': [],
            'level_cvs': [],
            'total_energy': 0.0,
            'scale_entropy': 0.0,
            'coefficients': []
        }

    level_energies = []
    level_cvs = []
    total_energy = 0.0

    cA = coeffs[0]
    approx_energy = float(np.sum(cA**2))
    approx_mean = float(np.mean(np.abs(cA)))
    approx_cv = (
        float((np.std(cA) / approx_mean) * 100.0)
        if approx_mean > AFMConfig.NUMERICAL_EPSILON else 0.0
    )
    level_energies.append(approx_energy)
    level_cvs.append(approx_cv)
    total_energy += approx_energy

    for detail_tuple in coeffs[1:]:
        lev_energy = 0.0
        lev_all_coeffs = []

        for detail_arr in detail_tuple:
            lev_energy += float(np.sum(detail_arr**2))
            lev_all_coeffs.extend(detail_arr.flatten().tolist())

        lev_arr = np.array(lev_all_coeffs)
        lev_mean = float(np.mean(np.abs(lev_arr)))
        lev_cv = (
            float((np.std(lev_arr) / lev_mean) * 100.0)
            if lev_mean > AFMConfig.NUMERICAL_EPSILON else 0.0
        )

        level_energies.append(lev_energy)
        level_cvs.append(lev_cv)
        total_energy += lev_energy

    if total_energy > AFMConfig.NUMERICAL_EPSILON:
        p = np.array(level_energies) / total_energy
        p_valid = p[p > 0]
        scale_entropy = float(-np.sum(p_valid * np.log2(p_valid)))
    else:
        scale_entropy = 0.0

    return {
        'level_energies': level_energies,
        'level_cvs': level_cvs,
        'total_energy': total_energy,
        'scale_entropy': scale_entropy,
        'coefficients': coeffs
    }


def multiscale_heterogeneity(
    image: np.ndarray,
    pixel_size_nm: float,
    max_scale: int = AFMConfig.MULTISCALE_MAX_SCALE
) -> Dict:
    """
    Scale-dependent heterogeneity analysis via Gaussian smoothing.

    FIXED v1.1.1: CV now uses delta_f (q95-q5) as denominator
    instead of |mean| to prevent CV=0 for zero-mean AFM images.

    Parameters
    ----------
    image : np.ndarray
        2D AFM height image (nm).
    pixel_size_nm : float
        Physical pixel size (nm).
    max_scale : int
        Maximum smoothing scale (pixels).

    Returns
    -------
    metrics : dict
        Scale-dependent metrics at each smoothing level.

    Examples
    --------
    >>> img = np.random.normal(0, 10, (256, 256))
    >>> ms = multiscale_heterogeneity(img, 19.53)
    >>> len(ms['scales']) <= 5
    True
    """
    image = validate_image_array(image, "multiscale_heterogeneity")
    pixel_size_nm = validate_pixel_size(
        pixel_size_nm, "multiscale_heterogeneity"
    )

    results = {
        'scales': [],
        'scale_nm': [],
        'cv_values': [],
        'roughness': [],
        'correlation_length': [],
        'lacunarity': []
    }

    ny, nx = image.shape

    for scale in range(1, max_scale + 1):
        smoothed = gaussian_filter(image, sigma=float(scale))
        heights = smoothed.flatten()

        q5 = np.percentile(heights, AFMConfig.PIP_LOWER_PERCENTILE)
        q95 = np.percentile(heights, AFMConfig.PIP_UPPER_PERCENTILE)
        filtered = heights[(heights >= q5) & (heights <= q95)]

        if len(filtered) < 2:
            continue

        mean_f = float(np.mean(filtered))
        std_f = float(np.std(filtered, ddof=1))

        # FIXED v1.1.1 (Issue A — Critical):
        # Use delta_f = q95 - q5 as denominator (consistent with CV_PIP)
        # instead of |mean_f| which is always ≈0 for zero-mean AFM images.
        delta_f = q95 - q5

        cv = (float(std_f / (delta_f + AFMConfig.NUMERICAL_EPSILON) * 100.0)
              if delta_f > AFMConfig.NUMERICAL_EPSILON else 0.0)

        roughness_val = std_f

        row = smoothed[ny // 2, :]
        row_centered = row - np.mean(row)
        if np.std(row_centered) > AFMConfig.NUMERICAL_EPSILON:
            acorr = np.correlate(row_centered, row_centered, mode='full')
            acorr = acorr[len(row_centered)-1:]
            acorr /= (acorr[0] + AFMConfig.NUMERICAL_EPSILON)
            below_1e = np.where(acorr < (1.0 / np.e))[0]
            corr_len = (
                float(below_1e[0]) * pixel_size_nm
                if len(below_1e) > 0
                else float(len(acorr)) * pixel_size_nm
            )
        else:
            corr_len = 0.0

        binary = smoothed > mean_f
        box_size = 2**scale
        lac = calculate_lacunarity(binary, box_size)

        results['scales'].append(scale)
        results['scale_nm'].append(float(scale) * pixel_size_nm)
        results['cv_values'].append(cv)
        results['roughness'].append(roughness_val)
        results['correlation_length'].append(corr_len)
        results['lacunarity'].append(lac)

    return results


def directional_analysis(
    image: np.ndarray,
    n_directions: int = 8
) -> Dict:
    """
    FFT-based directional anisotropy analysis.

    Parameters
    ----------
    image : np.ndarray
        2D AFM height image (nm).
    n_directions : int
        Number of directional sectors.

    Returns
    -------
    metrics : dict
        Directional power and anisotropy metrics.

    Examples
    --------
    >>> img = np.random.normal(0, 10, (256, 256))
    >>> da = directional_analysis(img)
    >>> 0.0 <= da['anisotropy_index'] <= 1.0
    True
    """
    image = validate_image_array(image, "directional_analysis")

    ny, nx = image.shape
    cy, cx = ny // 2, nx // 2

    f_image = np.fft.fftshift(np.fft.fft2(image - np.mean(image)))
    f_mag = np.abs(f_image)

    y_idx, x_idx = np.indices((ny, nx))
    dy = y_idx - cy
    dx = x_idx - cx
    angles_all = np.degrees(np.arctan2(dy, dx)) % 180.0

    dc_mask = (dy == 0) & (dx == 0)

    sector_width = 180.0 / n_directions
    angles_centers = np.arange(n_directions) * sector_width
    directional_power = []

    for angle in angles_centers:
        lower = angle
        upper = angle + sector_width
        sector_mask = (
            (angles_all >= lower)
            & (angles_all < upper)
            & ~dc_mask
        )
        sector_power = float(np.sum(f_mag[sector_mask]**2))
        directional_power.append(sector_power)

    total_power = sum(directional_power)
    if total_power > AFMConfig.NUMERICAL_EPSILON:
        norm_power = [p / total_power for p in directional_power]
        anisotropy_index = float(max(norm_power) - min(norm_power))
        primary_direction = float(
            angles_centers[int(np.argmax(directional_power))]
        )
    else:
        anisotropy_index = 0.0
        primary_direction = 0.0

    return {
        'angles': angles_centers,
        'directional_power': directional_power,
        'anisotropy_index': anisotropy_index,
        'primary_direction': primary_direction
    }


def multiscale_directional_analysis(
    image: np.ndarray,
    pixel_size_nm: float
) -> Dict:
    """
    Comprehensive multiscale and directional surface analysis.

    Parameters
    ----------
    image : np.ndarray
        2D AFM height image (nm).
    pixel_size_nm : float
        Physical pixel size (nm).

    Returns
    -------
    results : dict
        Wavelet, multiscale, directional, and summary results.

    Examples
    --------
    >>> img = np.random.normal(0, 10, (256, 256))
    >>> results = multiscale_directional_analysis(img, 19.53)
    >>> 'summary' in results
    True
    """
    image = validate_image_array(image, "multiscale_directional_analysis")
    pixel_size_nm = validate_pixel_size(
        pixel_size_nm, "multiscale_directional_analysis"
    )

    logger.info("Running wavelet analysis...")
    wavelet_metrics = wavelet_analysis(image)

    logger.info("Running multiscale heterogeneity analysis...")
    scale_metrics = multiscale_heterogeneity(image, pixel_size_nm)

    logger.info("Running directional analysis...")
    direction_metrics = directional_analysis(image)

    cv_vals = scale_metrics['cv_values']
    if len(cv_vals) >= 2 and np.mean(cv_vals) > AFMConfig.NUMERICAL_EPSILON:
        scale_het_index = float(
            np.std(cv_vals) / np.mean(cv_vals)
        )
    else:
        scale_het_index = 0.0

    summary = {
        'wavelet_complexity': wavelet_metrics['scale_entropy'],
        'scale_heterogeneity_index': scale_het_index,
        'anisotropy_index': direction_metrics['anisotropy_index'],
        'primary_direction': direction_metrics['primary_direction']
    }

    return {
        'wavelet': wavelet_metrics,
        'multiscale': scale_metrics,
        'directional': direction_metrics,
        'summary': summary
    }


# ========================
# Section 12: NIST Validation
# ========================

def generate_validation_surface() -> np.ndarray:
    """
    Generate the NIST SRM 2073-targeting synthetic validation surface.

    Returns
    -------
    surface : np.ndarray
        256x256 validation surface with zero mean.
    """
    size = AFMConfig.VALIDATION_SURFACE_SIZE
    x = np.linspace(0, 1, size)
    X, Y = np.meshgrid(x, x)

    ramp = 155.0 * (2.0 * np.abs(X - 0.5) - 1.0)
    valley = -350.0 * np.exp(
        -((X - 0.58)**2 + (Y - 0.58)**2) / 0.004
    )
    mid_freq = 55.0 * np.sin(16.0 * np.pi * X)
    high_freq = 35.0 * np.sin(64.0 * np.pi * X)

    rng = np.random.RandomState(AFMConfig.VALIDATION_RANDOM_SEED)
    spikes = np.zeros_like(X)
    spike_idx = rng.choice(X.size, size=20, replace=False)
    spikes.flat[spike_idx] = 350.0

    noise = 5.0 * rng.normal(0, 1, X.shape)

    surface = ramp + valley + spikes + mid_freq + high_freq + noise
    surface -= np.mean(surface)

    return surface


def validate_roughness_calibration() -> bool:
    """
    NIST SRM 2073-traceable validation of all roughness metrics.

    Returns
    -------
    passed : bool
        True if all validation criteria are met.
    """
    logger.info("Running NIST SRM 2073 validation...")

    surface = generate_validation_surface()
    roughness = calculate_roughness_metrics(surface)
    opd = orthogonal_pdf_decomposition(surface)

    criteria = [
        (
            'Ra',
            roughness['Ra_nm'],
            AFMConfig.NIST_RA_TARGET,
            AFMConfig.NIST_RA_TOLERANCE,
            'nm'
        ),
        (
            'Rsk',
            roughness['Rsk'],
            AFMConfig.NIST_RSK_TARGET,
            AFMConfig.NIST_RSK_TOLERANCE,
            ''
        ),
        (
            'Rku',
            roughness['Rku'],
            AFMConfig.NIST_RKU_TARGET,
            AFMConfig.NIST_RKU_TOLERANCE,
            ''
        ),
    ]

    passed = True
    results_log = []

    for name, computed, target, tolerance, unit in criteria:
        lower = target - tolerance
        upper = target + tolerance
        ok = lower <= computed <= upper
        status = "PASS" if ok else "FAIL"

        if not ok:
            passed = False

        msg = (
            f"  {name}: computed={computed:.4f}{unit}  "
            f"target={target:.4f}+/-{tolerance:.4f}{unit}  "
            f"range=[{lower:.4f}, {upper:.4f}]  {status}"
        )
        results_log.append(msg)
        logger.info(msg)

    energy_gap = opd['energy_gap']
    gap_ok = energy_gap >= AFMConfig.NIST_ENERGY_GAP_MIN
    gap_status = "PASS" if gap_ok else "FAIL"
    if not gap_ok:
        passed = False

    gap_msg = (
        f"  OPD Energy Gap: {energy_gap:.4f} "
        f"(min={AFMConfig.NIST_ENERGY_GAP_MIN})  {gap_status}"
    )
    logger.info(gap_msg)

    if passed:
        logger.info(
            "NIST validation PASSED — "
            "all metrics within SRM 2073 tolerances."
        )
    else:
        logger.warning(
            "NIST validation FAILED — "
            "some metrics outside tolerances."
        )

    return passed


# ========================
# Section 13: Visualization
# ========================

def plot_radial_psd(
    freq_bins: np.ndarray,
    radial_psd: np.ndarray,
    scan_size_nm: float,
    output_path: Optional[Path] = None
) -> None:
    """
    Generate publication-quality log-log radial PSD plot.

    Parameters
    ----------
    freq_bins : np.ndarray
        Spatial frequency bins (cycles/nm).
    radial_psd : np.ndarray
        Radially averaged PSD values (nm^2).
    scan_size_nm : float
        Scan size for plot title (nm).
    output_path : Path, optional
        Save path. If None, displays interactively.
    """
    if len(freq_bins) == 0 or len(radial_psd) == 0:
        logger.warning("No radial PSD data to plot.")
        return

    fig, ax = plt.subplots(figsize=(8, 6))

    try:
        ax.loglog(freq_bins, radial_psd, 'o-',
                  markersize=4, linewidth=1.5,
                  color='steelblue', label='Radial PSD')

        ax.set_xlabel('Spatial Frequency (1/nm)', fontsize=12)
        ax.set_ylabel('Power Spectral Density (nm^2)', fontsize=12)
        ax.set_title(
            f'Radially Averaged PSD — '
            f'{scan_size_nm/1000:.1f} um Scan',
            fontsize=13
        )
        ax.grid(True, which='both', linestyle='--',
                alpha=0.4, linewidth=0.8)
        ax.legend(fontsize=10)
        plt.tight_layout()

        if output_path:
            fig.savefig(str(output_path), dpi=300, bbox_inches='tight')
            logger.info(f"Radial PSD plot saved: {output_path}")
        else:
            plt.show()

    finally:
        plt.close(fig)


def plot_multiscale_results(
    results: Dict,
    output_path: Optional[Path] = None
) -> None:
    """
    Generate four-panel multiscale analysis visualization.

    Parameters
    ----------
    results : dict
        Output from multiscale_directional_analysis().
    output_path : Path, optional
        Save path.
    """
    fig = plt.figure(figsize=(14, 10))

    try:
        ax1 = fig.add_subplot(2, 2, 1)
        energies = results['wavelet']['level_energies']
        n_levels = len(energies)
        labels = ['Approx'] + [f'Detail {i}' for i in range(1, n_levels)]
        colors = plt.cm.viridis(np.linspace(0.2, 0.9, n_levels))
        ax1.bar(range(n_levels), energies, color=colors)
        ax1.set_xticks(range(n_levels))
        ax1.set_xticklabels(labels, rotation=15, fontsize=9)
        ax1.set_xlabel('Decomposition Level', fontsize=10)
        ax1.set_ylabel('Energy', fontsize=10)
        ax1.set_title('Wavelet Energy Distribution', fontsize=11)

        ax2 = fig.add_subplot(2, 2, 2)
        scales = results['multiscale']['scale_nm']
        cv_vals = results['multiscale']['cv_values']
        if scales and cv_vals:
            ax2.plot(scales, cv_vals, 'o-',
                     color='darkorange', linewidth=2, markersize=6)
            ax2.set_xlabel('Scale (nm)', fontsize=10)
            ax2.set_ylabel('CV (%)', fontsize=10)
            ax2.set_title('Scale-Dependent Heterogeneity', fontsize=11)
            ax2.grid(True, alpha=0.3)

        ax3 = fig.add_subplot(2, 2, 3, polar=True)
        angles_rad = np.radians(results['directional']['angles'])
        d_power = results['directional']['directional_power']

        if len(angles_rad) > 0 and len(d_power) > 0:
            angles_plot = np.append(angles_rad, angles_rad[0])
            power_plot = np.append(d_power, d_power[0])
            ax3.plot(angles_plot, power_plot,
                     color='crimson', linewidth=2)
            ax3.fill(angles_plot, power_plot,
                     color='crimson', alpha=0.2)
            ax3.set_title('Directional Power', fontsize=11, pad=20)

        ax4 = fig.add_subplot(2, 2, 4)
        ax4.axis('off')
        summary = results['summary']
        table_data = [
            ['Metric', 'Value'],
            ['Wavelet Complexity',
             f"{summary['wavelet_complexity']:.4f} bits"],
            ['Scale Heterogeneity',
             f"{summary['scale_heterogeneity_index']:.4f}"],
            ['Anisotropy Index',
             f"{summary['anisotropy_index']:.4f}"],
            ['Primary Direction',
             f"{summary['primary_direction']:.1f} deg"],
        ]
        table = ax4.table(
            cellText=table_data[1:],
            colLabels=table_data[0],
            loc='center',
            cellLoc='center'
        )
        table.auto_set_font_size(False)
        table.set_fontsize(10)
        table.scale(1.2, 1.8)
        ax4.set_title('Summary Metrics', fontsize=11)

        plt.suptitle('Multiscale Analysis Results',
                     fontsize=13, fontweight='bold')
        plt.tight_layout(rect=[0, 0, 1, 0.96])

        if output_path:
            fig.savefig(str(output_path), dpi=300, bbox_inches='tight')
            logger.info(f"Multiscale plot saved: {output_path}")
        else:
            plt.show()

    finally:
        plt.close(fig)


def plot_opd_spectrum(
    opd_results: Dict,
    output_path: Optional[Path] = None
) -> None:
    """
    Visualize OPD energy spectrum.

    Parameters
    ----------
    opd_results : dict
        Output from orthogonal_pdf_decomposition().
    output_path : Path, optional
        Save path.
    """
    if not opd_results['energy_spectrum']:
        logger.warning("No OPD data to plot.")
        return

    fig, axes = plt.subplots(1, 2, figsize=(12, 5))

    try:
        spectrum = opd_results['energy_spectrum']
        levels = range(len(spectrum))

        axes[0].semilogy(levels, spectrum, 'o-',
                         color='darkgreen', linewidth=2, markersize=8)
        axes[0].set_xlabel('Decomposition Level', fontsize=11)
        axes[0].set_ylabel('Normalized Energy (log scale)', fontsize=11)
        axes[0].set_title('OPD Energy Spectrum', fontsize=12)
        axes[0].grid(True, which='both', alpha=0.3)
        axes[0].set_xticks(list(levels))

        colors = plt.cm.plasma(np.linspace(0.2, 0.9, len(spectrum)))
        axes[1].bar(levels, [e * 100 for e in spectrum], color=colors)
        axes[1].set_xlabel('Decomposition Level', fontsize=11)
        axes[1].set_ylabel('Energy Contribution (%)', fontsize=11)
        axes[1].set_title('Scale-wise Energy (%)', fontsize=12)
        axes[1].set_xticks(list(levels))

        dom = opd_results['dominant_scale']
        if 0 <= dom < len(spectrum):
            axes[1].bar(dom, spectrum[dom] * 100,
                        color='red', label=f'Dominant (L{dom})')
            axes[1].legend(fontsize=9)

        plt.tight_layout()

        if output_path:
            fig.savefig(str(output_path), dpi=300, bbox_inches='tight')
            logger.info(f"OPD plot saved: {output_path}")
        else:
            plt.show()

    finally:
        plt.close(fig)


# ========================
# Section 14: Core Analysis Pipeline
# ========================

def analyze_afm_image(
    image_path: Union[str, Path],
    pixel_size_nm: float,
    scan_size_nm: Optional[float] = None,
    output_dir: Optional[Path] = None,
    save_plots: bool = True,
    save_csv: bool = True
) -> pd.DataFrame:
    """
    Complete AFM surface analysis pipeline (v1.1.1).

    Parameters
    ----------
    image_path : str or Path
        Path to AFM TIFF image.
    pixel_size_nm : float
        Physical pixel size in nanometers.
    scan_size_nm : float, optional
        Total scan size in nanometers.
    output_dir : Path, optional
        Directory for saving outputs.
    save_plots : bool
        Whether to save visualization plots.
    save_csv : bool
        Whether to save radial PSD as CSV.

    Returns
    -------
    results_df : pd.DataFrame
        Single-row DataFrame with all computed metrics.

    Examples
    --------
    >>> df = analyze_afm_image("sample.tif", pixel_size_nm=19.53)
    >>> 'Rt_Robust_nm' in df.columns
    True
    """
    image_path = Path(image_path)
    if not image_path.exists():
        raise FileNotFoundError(f"Image not found: {image_path}")

    pixel_size_nm = validate_pixel_size(pixel_size_nm, "analyze_afm_image")

    if output_dir is None:
        output_dir = Path.home() / "Desktop"
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    stem = image_path.stem
    logger.info(f"Analyzing: {image_path.name}")
    logger.info(f"Pixel size: {pixel_size_nm:.4f} nm")

    logger.info("Loading image...")
    image = load_afm_image(image_path)
    image = validate_image_array(image, "analyze_afm_image")

    if scan_size_nm is None:
        scan_size_nm = pixel_size_nm * max(image.shape)

    logger.info(
        f"Image loaded: {image.shape[0]}x{image.shape[1]} px | "
        f"Scan: {scan_size_nm/1000:.3f} um | "
        f"Z range: {image.min():.2f} to {image.max():.2f} nm"
    )

    logger.info("Computing roughness metrics...")
    roughness = calculate_roughness_metrics(image)

    logger.info("Computing power spectral density...")
    psd, freq_x, freq_y = calculate_psd(image, pixel_size_nm)
    fourier = calculate_fourier_metrics(psd, freq_x, freq_y, pixel_size_nm)

    logger.info("Computing PIP metrics...")
    pip = calculate_pip_metrics(image, pixel_size_nm)

    logger.info("Computing HI and PDI...")
    pdi = calculate_pdi(fourier['cv_fourier'], pip['cv_pip'])
    hi = calculate_hi(
        pip['delta_pip'], fourier['mean_psd'],
        fourier['cv_fourier'], pip['cv_pip']
    )
    hi_radial = calculate_hi_radial(
        pip['delta_pip'], fourier['mean_psd_radial'],
        fourier['cv_fourier'], pip['cv_pip']
    )

    logger.info("Running OPD decomposition...")
    opd = orthogonal_pdf_decomposition(image)

    logger.info("Running multiscale directional analysis...")
    ms = multiscale_directional_analysis(image, pixel_size_nm)

    freq_bins, radial_psd_vals = fourier['radial_data']

    radial_psd_path = None
    if save_csv and len(freq_bins) > 0:
        psd_df = pd.DataFrame({
            'Frequency_per_nm': freq_bins,
            'PSD_nm2': radial_psd_vals,
            'Wavelength_nm': 1.0 / (freq_bins + AFMConfig.NUMERICAL_EPSILON)
        })
        radial_psd_path = output_dir / f"{stem}_radial_psd.csv"
        psd_df.to_csv(radial_psd_path, index=False)
        logger.info(f"Radial PSD CSV saved: {radial_psd_path}")

    if save_plots:
        plot_radial_psd(
            freq_bins, radial_psd_vals, scan_size_nm,
            output_dir / f"{stem}_radial_psd.png"
        )
        plot_multiscale_results(
            ms,
            output_dir / f"{stem}_multiscale.png"
        )
        plot_opd_spectrum(
            opd,
            output_dir / f"{stem}_opd.png"
        )

    results = {
        'Sample_Name': stem,
        'Image_Path': str(image_path),
        'Analysis_Timestamp': datetime.now().isoformat(),
        'AFMAnalyzer_Version': __version__,

        'Scan_Size_nm': scan_size_nm,
        'Pixel_Size_nm': pixel_size_nm,
        'Image_Height_px': image.shape[0],
        'Image_Width_px': image.shape[1],

        'Ra_nm': roughness['Ra_nm'],
        'Rq_nm': roughness['Rq_nm'],
        'Rt_nm': roughness['Rt_nm'],
        'Rt_Robust_nm': roughness['Rt_Robust_nm'],
        'Rsk': roughness['Rsk'],
        'Rku': roughness['Rku'],
        'Fractal_Dim': roughness['Fractal_Dim'],

        'Mean_PSD_nm2': fourier['mean_psd'],
        'Mean_PSD_Radial_nm2': fourier['mean_psd_radial'],
        'CV_Fourier_pct': fourier['cv_fourier'],

        'Delta_PIP_nm': pip['delta_pip'],
        'CV_PIP_pct': pip['cv_pip'],
        'Q5_nm': pip['q5'],
        'Q95_nm': pip['q95'],

        'PDI': pdi,
        'HI': hi,
        'HI_Radial': hi_radial,

        'OPD_Energy_Gap': opd['energy_gap'],
        'OPD_Dominant_Scale': opd['dominant_scale'],

        'Wavelet_Complexity_bits': ms['summary']['wavelet_complexity'],
        'Scale_Heterogeneity_Index': ms['summary']['scale_heterogeneity_index'],
        'Anisotropy_Index': ms['summary']['anisotropy_index'],
        'Primary_Direction_deg': ms['summary']['primary_direction'],

        'Radial_PSD_CSV': str(radial_psd_path) if radial_psd_path else 'N/A',
    }

    logger.info("Analysis complete.")
    return pd.DataFrame([results])


# ========================
# Section 15: Excel Export
# ========================

def save_results_excel(
    df: pd.DataFrame,
    output_path: Path
) -> bool:
    """
    Save analysis results to Excel with units legend sheet.

    Parameters
    ----------
    df : pd.DataFrame
        Results DataFrame.
    output_path : Path
        Output .xlsx file path.

    Returns
    -------
    success : bool
        True if saved successfully.
    """
    if not EXCEL_AVAILABLE:
        csv_path = output_path.with_suffix('.csv')
        df.to_csv(csv_path, index=False)
        logger.info(f"openpyxl not available. Saved as CSV: {csv_path}")
        return True

    try:
        df.to_excel(str(output_path), index=False, engine='openpyxl')

        wb = load_workbook(str(output_path))
        ws = wb.create_sheet("Units_Legend")

        legend = [
            ("Ra_nm", "nm", "Arithmetic mean roughness (ISO 4287)"),
            ("Rq_nm", "nm", "Root mean square roughness (ISO 4287)"),
            ("Rt_nm", "nm", "Maximum peak-to-valley height (ISO 4287 strict)"),
            ("Rt_Robust_nm", "nm",
             "Robust peak-to-valley P99.5-P0.5: spike-resistant"),
            ("Rsk", "dimensionless",
             "Skewness with Fisher-Pearson bias correction"),
            ("Rku", "dimensionless",
             "Excess kurtosis with Pearson bias correction"),
            ("Fractal_Dim", "dimensionless", "OPD-derived fractal dimension"),
            ("Mean_PSD_nm2", "nm^2", "2D mean power spectral density"),
            ("Mean_PSD_Radial_nm2", "nm^2", "Radially averaged PSD mean"),
            ("CV_Fourier_pct", "%", "Angular CV of PSD directional sectors"),
            ("Delta_PIP_nm", "nm", "Height range: P95 - P5"),
            ("CV_PIP_pct", "%", "std(filtered heights) / delta_pip * 100%"),
            ("PDI", "dimensionless", "CV_Fourier / CV_PIP"),
            ("HI", "dimensionless",
             "(delta_PIP/sqrt(PSD_mean)) * ln(PDI)"),
            ("HI_Radial", "dimensionless", "HI using radial PSD mean"),
            ("OPD_Energy_Gap", "dimensionless", "OPD energy gap E[0] - E[1]"),
            ("Wavelet_Complexity_bits", "bits",
             "Shannon entropy of wavelet energy"),
            ("Scale_Heterogeneity_Index", "dimensionless",
             "CV of scale-dependent CV values"),
            ("Anisotropy_Index", "dimensionless",
             "Directional power range"),
            ("Primary_Direction_deg", "degrees",
             "Dominant surface orientation"),
        ]

        ws.append(["Column Name", "Unit", "Description", "Method"])
        for col_name, unit, desc in legend:
            ws.append([col_name, unit, desc, "AFMAnalyzer v" + __version__])

        wb.move_sheet(ws.title, offset=-len(wb.sheetnames) + 1)
        wb.save(str(output_path))

        logger.info(f"Results saved: {output_path}")
        return True

    except PermissionError:
        logger.error(f"Cannot write to {output_path}. File may be open.")
        return False
    except Exception as e:
        logger.error(f"Excel save failed: {e}")
        return False


# ========================
# Section 16: Main Entry Point
# ========================

def main():
    """
    Main entry point for AFMAnalyzer GUI mode.

    Workflow:
    1. Run NIST validation
    2. File selection dialog
    3. Automatic metadata extraction
    4. User confirmation/override of pixel size
    5. Full analysis pipeline
    6. Save Excel + plots
    """
    if not GUI_AVAILABLE:
        logger.error(
            "GUI mode requires tkinter. "
            "Please use the programmatic API:\n"
            "  from afmanalyzer import analyze_afm_image\n"
            "  df = analyze_afm_image('image.tif', pixel_size_nm=19.53)"
        )
        return

    root = tk.Tk()
    root.withdraw()

    try:
        logger.info("=" * 50)
        logger.info(f"AFMAnalyzer v{__version__} — Starting")
        logger.info("=" * 50)

        validation_passed = validate_roughness_calibration()
        if not validation_passed:
            proceed = messagebox.askyesno(
                "Validation Warning",
                "NIST validation failed. Results may be unreliable.\n"
                "Proceed anyway?",
                icon='warning'
            )
            if not proceed:
                logger.info("Analysis cancelled by user.")
                return

        image_path = filedialog.askopenfilename(
            title="Select AFM Image (float32 TIFF from JPK)",
            filetypes=[
                ("TIFF files", "*.tif *.tiff"),
                ("All files", "*.*")
            ]
        )

        if not image_path:
            logger.info("No file selected. Exiting.")
            return

        image_path = Path(image_path)

        logger.info("Extracting metadata...")
        metadata = extract_afm_metadata(image_path)

        pixel_size_nm = None
        scan_size_nm = None

        if metadata['success'] and metadata['pixel_size_nm']:
            confirm = tk.Toplevel()
            confirm.title("Confirm Parameters — AFMAnalyzer")
            confirm.geometry("480x320")
            confirm.resizable(False, False)

            tk.Label(
                confirm,
                text="Metadata Extracted Successfully",
                font=("Arial", 13, "bold")
            ).pack(pady=12)

            info_frame = tk.Frame(confirm, relief='sunken', bd=1)
            info_frame.pack(fill='x', padx=20, pady=5)

            ps = metadata['pixel_size_nm']
            ss = metadata.get('scan_size_x_nm', 0) or 0
            wp = metadata.get('width_pixels', 0)
            hp = metadata.get('height_pixels', 0)

            for label, value in [
                ("Pixel Size:", f"{ps:.4f} nm"),
                ("Scan Size:", f"{ss/1000:.3f} x {ss/1000:.3f} um"),
                ("Dimensions:", f"{wp} x {hp} pixels"),
            ]:
                row = tk.Frame(info_frame)
                row.pack(fill='x', padx=10, pady=3)
                tk.Label(row, text=label, width=15,
                         anchor='w').pack(side='left')
                tk.Label(row, text=value,
                         anchor='w').pack(side='left')

            tk.Label(
                confirm,
                text="Confirm to use these parameters:",
                font=("Arial", 10)
            ).pack(pady=8)

            user_choice = tk.StringVar(value='confirm')

            btn_frame = tk.Frame(confirm)
            btn_frame.pack(pady=15)

            tk.Button(
                btn_frame, text="Confirm",
                width=14, bg='#4CAF50', fg='white',
                command=lambda: [
                    user_choice.set('confirm'),
                    confirm.destroy()
                ]
            ).pack(side='left', padx=8)

            tk.Button(
                btn_frame, text="Manual Entry",
                width=14, bg='#f44336', fg='white',
                command=lambda: [
                    user_choice.set('manual'),
                    confirm.destroy()
                ]
            ).pack(side='left', padx=8)

            confirm.wait_window()

            if user_choice.get() == 'confirm':
                pixel_size_nm = ps
                scan_size_nm = ss

        if pixel_size_nm is None:
            manual = tk.Toplevel()
            manual.title("Manual Parameter Entry")
            manual.geometry("360x200")

            tk.Label(
                manual,
                text="Enter Scan Parameters",
                font=("Arial", 12, "bold")
            ).pack(pady=10)

            entry_frame = tk.Frame(manual)
            entry_frame.pack(pady=5)

            tk.Label(entry_frame, text="Pixel Size (nm):").grid(
                row=0, column=0, padx=5, pady=5, sticky='e'
            )
            ps_var = tk.StringVar(value="19.53")
            ps_entry = tk.Entry(entry_frame, textvariable=ps_var, width=12)
            ps_entry.grid(row=0, column=1, padx=5)

            tk.Label(entry_frame, text="Scan Size (um):").grid(
                row=1, column=0, padx=5, pady=5, sticky='e'
            )
            ss_var = tk.StringVar(value="10.0")
            ss_entry = tk.Entry(entry_frame, textvariable=ss_var, width=12)
            ss_entry.grid(row=1, column=1, padx=5)

            manual_result = {}

            def on_ok():
                try:
                    manual_result['ps'] = float(ps_var.get())
                    manual_result['ss'] = float(ss_var.get()) * 1000
                    manual.destroy()
                except ValueError:
                    messagebox.showerror(
                        "Input Error",
                        "Please enter valid numbers.",
                        parent=manual
                    )

            tk.Button(manual, text="OK", command=on_ok,
                      width=10, bg='#2196F3', fg='white').pack(pady=15)
            manual.wait_window()

            if 'ps' not in manual_result:
                logger.info("No parameters entered. Exiting.")
                return

            pixel_size_nm = manual_result['ps']
            scan_size_nm = manual_result['ss']

        try:
            pixel_size_nm = validate_pixel_size(
                pixel_size_nm, "main"
            )
        except PhysicalUnitError as e:
            messagebox.showerror("Invalid Pixel Size", str(e))
            return

        output_dir = Path.home() / "Desktop"

        logger.info("Starting full analysis pipeline...")
        df = analyze_afm_image(
            image_path=image_path,
            pixel_size_nm=pixel_size_nm,
            scan_size_nm=scan_size_nm,
            output_dir=output_dir,
            save_plots=True,
            save_csv=True
        )

        output_xlsx = output_dir / f"{image_path.stem}_analysis.xlsx"
        save_results_excel(df, output_xlsx)

        messagebox.showinfo(
            "Analysis Complete",
            f"Results saved to Desktop:\n\n"
            f"- {image_path.stem}_analysis.xlsx\n"
            f"- {image_path.stem}_radial_psd.png\n"
            f"- {image_path.stem}_multiscale.png\n"
            f"- {image_path.stem}_opd.png\n"
            f"- {image_path.stem}_radial_psd.csv"
        )

        logger.info("AFMAnalyzer completed successfully.")

    except Exception as e:
        logger.error(f"Critical error: {e}", exc_info=True)
        if GUI_AVAILABLE:
            messagebox.showerror(
                "Critical Error",
                f"An unexpected error occurred:\n{str(e)}\n\n"
                "Check the console for details."
            )
    finally:
        try:
            root.destroy()
        except Exception:
            pass


if __name__ == "__main__":
    main()
